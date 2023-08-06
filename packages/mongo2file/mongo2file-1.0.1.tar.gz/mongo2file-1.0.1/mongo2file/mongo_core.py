# -*- coding:utf-8 -*-
import math
import os
import warnings
from typing import Optional
import pyarrow.csv as pa_csv_
import pyarrow.feather as pa_feather_
import pyarrow.parquet as pa_parquet_
import pyarrow as pa
import xlsxwriter
from alive_progress import alive_bar
from colorama import init as colorama_init_, Fore
from dotenv import load_dotenv
from pymongo import MongoClient

from .constants import *
from .utils import to_str_datetime, serialize_obj, csv_concurrent_, excel_concurrent_, concurrent_, \
    json_concurrent_, schema_, no_collection_to_csv_, no_collection_to_excel_, no_collection_to_json_

load_dotenv(verbose=True)
colorama_init_(autoreset=True)


def check_folder_path(folder_path):
    if folder_path is None:
        _ = '.'
    elif not os.path.exists(folder_path):
        os.makedirs(folder_path)
        _ = folder_path
    else:
        _ = folder_path
    return _


class MongoEngine:

    def __init__(
            self,
            host: Optional[str] = None,
            port: Optional[int] = None,
            username: Optional[str] = None,
            password: Optional[str] = None,
            database: Optional[str] = None,
            collection: Optional[str] = None,
            conn_timeout: Optional[int] = MONGO_CONN_TIMEOUT,  # noqa F401
            conn_retries: Optional[int] = MONGO_CONN_RETRIES  # noqa F401
    ):
        self.host = MONGO_HOST if host is None else host
        self.port = MONGO_PORT if port is None else port
        self.username = username
        self.password = password
        self.database = MONGO_DATABASE if database is None else database
        self.collection = MONGO_COLLECTION if collection is None else collection
        self.conn_timeout = MONGO_CONN_TIMEOUT if conn_timeout is None else conn_timeout
        self.conn_retries = MONGO_CONN_RETRIES if conn_retries is None else conn_retries
        self.mongo_core_ = MongoClient(
            host=self.host,
            port=self.port,
            username=self.username,
            password=self.password,
            maxPoolSize=MONGO_POOL_MAX_WORKERS
        )
        self.db_ = self.mongo_core_[self.database]
        self.collection_names = self.db_.list_collection_names()
        if self.collection:
            self.collection_ = self.db_[self.collection]
        else:
            self.collection_ = None

    def to_csv(self, query=None, folder_path: str = None, filename: str = None, _id: bool = False, limit: int = -1,
               is_block: bool = False, block_size: int = 10000, ignore_error: bool = False):
        """
        :param query: 数据库查询条件、字典类型、只作用于单表导出
        :param folder_path: 指定导出的目录
        :param filename: 指定导出的文件名
        :param _id: 是否导出 _id 默认否
        :param limit: 限制数据表查询的条数
        :param is_block: 是否分块导出
        :param block_size: 块大小、is_block 为 True 时生效
        :param ignore_error: 是否忽略错误、数据表中存在非序列化类型时使用、这将影响程序的性能
        """
        if query is None:
            query = {}
        if not isinstance(query, dict):
            raise TypeError('query must be of dict type')
        if not isinstance(limit, int):
            raise TypeError("limit must be an integer type")
        if not isinstance(_id, bool):
            raise TypeError("_id must be an boolean type")
        folder_path_ = check_folder_path(folder_path)

        if self.collection_:
            """
            stats_ = self.db_.command('collstats', self.collection)
            print(f'命名空间: {stats_.get("ns")}, '
                  f'内存总大小: {round(stats_.get("size") / 1024, 2)} KB, '
                  f'存储大小: {round(stats_.get("storageSize") / 1024, 2)} KB, '
                  f'对象平均大小: {round(stats_.get("avgObjSize") / 1024, 2)} KB, '
                  f'文档数: {stats_.get("count")}')
            """

            if filename is None:
                filename = f'{self.collection}_{to_str_datetime()}.csv'

            if is_block:
                count_ = self.collection_.count_documents(query)
                block_count_ = math.ceil(count_ / block_size)
                csv_concurrent_(self.save_csv_, self.collection, block_count_, block_size, folder_path_, ignore_error)
                result_ = ECHO_INFO.format(Fore.GREEN, self.collection, folder_path_)
            else:
                doc_list = self.collection_.find(query, {'_id': 0}).limit(limit) \
                    if limit != -1 else self.collection_.find(query, {'_id': 0})

                """
                todo polars
                [{"a":111,"b":222},{"a":111,"b":222}] -> [{"a":[111,111],"b":[222,222]}]
                doc_list_ = [list(doc.values()) for doc in doc_list]
                columns_ = list(doc_list[0].keys())
                data = pl.DataFrame(doc_list_,columns=columns_)
                data.to_csv(file=f'{folder_path_}/{filename}', has_header=True)

                todo pandas
                df_ = DataFrame(data=doc_list)
                df_.to_csv(path_or_buf=f'{folder_path_}/{filename}', index=False, encoding=PANDAS_ENCODING)
                """

                doc_list_ = [schema_(doc_) for doc_ in doc_list]
                # schema_ = pa.schema([
                #     ('city', pa.string()),
                #     ('content', pa.string()),
                #     ('scenic_id', pa.string()),
                #     ('scenic_name', pa.string()),
                #     ('username', pa.string()),
                #     # ('update_date', pa.string())
                # ])
                df_ = pa.Table.from_pylist(mapping=doc_list_, schema=None, metadata=None)
                with pa_csv_.CSVWriter(f'{folder_path_}/{filename}', df_.schema) as writer:
                    writer.write_table(df_)

                # title_ = f'{Fore.GREEN} {self.collection} → {folder_path_}/{filename}'
                # count_ = self.collection_.count_documents(query)
                # with alive_bar(count_, title=title_, bar="blocks") as bar:
                # with codecs.open(f'{folder_path_}/{filename}', 'w', 'utf-8') as csvfile:
                #     writer = csv.writer(csvfile)
                #     writer.writerow(list(dict(doc_list[0]).keys()))
                #     # writer.writerows([list(dict(data_).values()) for data_ in doc_list])
                #     for data_ in doc_list:
                #         writer.writerow(list(dict(data_).values()))
                #         bar()

                result_ = ECHO_INFO.format(Fore.GREEN, self.collection, f'{folder_path_}/{filename}')
            return result_
        else:
            warnings.warn('No collection specified, All collections will be exported.', DeprecationWarning)
            self.to_many_collection_(no_collection_to_csv_, folder_path_, ignore_error)
            result_ = ECHO_INFO.format(Fore.GREEN, self.database, folder_path_)
            return result_

    def to_excel(self, query=None, folder_path: str = None, filename: str = None, _id: bool = False, limit: int = -1,
                 is_block: bool = False, block_size: int = 10000, mode: str = 'xlsx', ignore_error: bool = False):
        """
        :param query: 数据库查询条件 字典类型
        :param folder_path: 指定导出的目录
        :param filename: 指定导出的文件名
        :param _id: 是否导出 _id 默认否
        :param limit: 限制数据表查询的条数
        :param is_block: 是否分块导出
        :param block_size: 块大小、is_block 为 True 时生效
        :param mode: 枚举类型、 sheet(子表) 或 xlsx、is_block 为 True 时生效
        :param ignore_error: 是否忽略错误、数据表中存在非序列化类型时使用、这将影响程序的性能
        :return:
        """
        if query is None:
            query = {}
        if not isinstance(query, dict):
            raise TypeError('query must be of dict type')
        if not isinstance(limit, int):
            raise TypeError("limit must be an integer type")
        if not isinstance(_id, bool):
            raise TypeError("_id must be an boolean type")
        folder_path_ = check_folder_path(folder_path)
        if self.collection_:
            if filename is None:
                filename = f'{self.collection}_.xlsx'
            if is_block:
                count_ = self.collection_.count_documents(query)
                block_count_ = math.ceil(count_ / block_size)
                if mode not in ['xlsx', 'sheet']:
                    raise ValueError("mode must be specified as xlsx or sheet")
                if mode == 'xlsx':
                    excel_concurrent_(self.save_excel_, None, self.collection, block_count_, block_size, folder_path_,
                                      ignore_error)
                    result_ = ECHO_INFO.format(Fore.GREEN, self.collection, folder_path_)
                else:
                    with xlsxwriter.Workbook(f'{folder_path_}/{filename}') as work_book_:
                        excel_concurrent_(self.save_excel_, work_book_, self.collection, block_count_, block_size,
                                          folder_path_, ignore_error)
                    result_ = ECHO_INFO.format(Fore.GREEN, self.collection, folder_path_)
            else:
                doc_objs_ = self.collection_.find(query, {'_id': 0}).limit(
                    limit) if limit != -1 else self.collection_.find(query, {'_id': 0})

                """
                todo xlwings
                import xlwings as xw
                app = xw.App(visible=False, add_book=False)
                wb = app.books.open(f'{folder_path_}/{filename}') # 打开Excel文件
                sheet = wb.sheets[0]
                for index, doc in enumerate(doc_objs_):
                    sheet.range(f'A{index+1}').value = list(dict(doc).values())
                wb.save()
                wb.close()

                todo openpyxl
                from openpyxl import Workbook
                wb = Workbook(write_only=True)
                ws = wb.create_sheet(title="hi")
                for doc in doc_objs_:
                    ws.append(list(dict(doc).values()))
                wb.save(filename)
                wb.close()
                """

                """
                doc_list_ = [schema_(doc_) for doc_ in doc_objs_]
                df_ = pa.Table.from_pylist(mapping=doc_list_, schema=None)
                df_.to_pandas().to_excel(excel_writer=f'{folder_path}/{filename}.xlsx', index=False, encoding=PANDAS_ENCODING)
                """

                with xlsxwriter.Workbook(f'{folder_path_}/{filename}') as work_book_:
                    work_sheet_ = work_book_.add_worksheet('Sheet1')
                    title_ = f'{Fore.GREEN} {self.collection} → {folder_path_}/{filename}'
                    count_ = self.collection_.count_documents(query)
                    with alive_bar(count_, title=title_, bar="blocks") as bar:
                        header_ = list(dict(doc_objs_[0]).keys())
                        work_sheet_.write_row(f"A1", header_)
                        if ignore_error:
                            for index_, doc_ in enumerate(doc_objs_):
                                write_list_ = [
                                    doc_.get(x_) if doc_.get(x_) and isinstance(doc_.get(x_), IGNORE_TYPE) else None
                                    for x_ in header_]
                                work_sheet_.write_row(f"A{index_ + 2}", write_list_)
                                bar()
                        else:
                            for index_, doc in enumerate(doc_objs_):
                                write_list_ = [doc_ if isinstance(doc_, (int, float)) or doc_ is None else str(doc_) for
                                               doc_ in
                                               dict(doc).values()]
                                work_sheet_.write_row(f"A{index_ + 2}", write_list_)
                                bar()

                result_ = ECHO_INFO.format(Fore.GREEN, self.collection, f'{folder_path_}/{filename}')
            return result_
        else:
            warnings.warn('No collection specified, All collections will be exported.', DeprecationWarning)
            self.to_many_collection_(no_collection_to_excel_, folder_path_, ignore_error)
            result_ = ECHO_INFO.format(Fore.GREEN, self.database, folder_path_)
            return result_

    def to_json(self, query=None, folder_path: str = None, filename: str = None, _id: bool = False, limit: int = -1,
                is_block: bool = False, block_size: int = 10000, ignore_error: bool = False):
        """
        :param query: 数据库查询条件、字典类型、只作用于单表导出
        :param folder_path: 指定导出的目录
        :param filename: 指定导出的文件名
        :param _id: 是否导出 _id 默认否
        :param limit: 限制数据表查询的条数
        :param is_block: 是否分块导出
        :param block_size: 块大小、is_block 为 True 时生效
        :param ignore_error: 是否忽略错误、数据表中存在非序列化类型时使用、这将影响程序的性能
        :return:
        """
        if query is None:
            query = {}
        if not isinstance(query, dict):
            raise TypeError('query must be of dict type')
        if not isinstance(limit, int):
            raise TypeError("limit must be an integer type")
        if not isinstance(_id, bool):
            raise TypeError("_id must be an boolean type")
        folder_path_ = check_folder_path(folder_path)

        if self.collection_:
            if filename is None:
                filename = f'{self.collection}_{to_str_datetime()}.json'
                if is_block:
                    count_ = self.collection_.count_documents(query)
                    block_count_ = math.ceil(count_ / block_size)
                    json_concurrent_(self.save_json_, self.collection, block_count_, block_size, folder_path_,
                                     ignore_error)
                    result_ = ECHO_INFO.format(Fore.GREEN, self.collection, folder_path_)
                else:
                    doc_objs_ = self.collection_.find(query, {"_id": 0}).limit(
                        limit) if limit != -1 else self.collection_.find(query, {"_id": 0})
                    doc_list_ = list(doc_objs_)
                    data = {'RECORDS': doc_list_}
                    with open(f'{folder_path_}/{filename}', 'w', encoding="utf-8") as f:
                        f.write(serialize_obj(data))
                    result_ = ECHO_INFO.format(Fore.GREEN, self.collection, f'{folder_path_}/{filename}')
                return result_
        else:
            warnings.warn('No collection specified, All collections will be exported.', DeprecationWarning)
            self.to_many_collection_(no_collection_to_json_, folder_path_, ignore_error)
            result_ = ECHO_INFO.format(Fore.GREEN, self.database, folder_path_)
            return result_

    def to_pickle(self, query=None, folder_path: str = None, filename: str = None, _id: bool = False, limit: int = -1):
        if query is None:
            query = {}
        if not isinstance(query, dict):
            raise TypeError('query must be of dict type')
        if not isinstance(limit, int):
            raise TypeError("limit must be an integer type")
        if not isinstance(_id, int):
            raise TypeError("_id must be an boolean type")
        folder_path_ = check_folder_path(folder_path)

        if self.collection_:
            if filename is None:
                filename = f'{self.collection}_{to_str_datetime()}.pkl'
            doc_objs_ = self.collection_.find(query, {"_id": 0}).limit(
                limit) if limit != -1 else self.collection_.find(query, {"_id": 0})
            doc_list_ = [schema_(doc_) for doc_ in doc_objs_]

            import pickle
            with open(f'{folder_path_}/{filename}', 'wb') as f:
                pickle.dump(doc_list_, f)

            result_ = ECHO_INFO.format(Fore.GREEN, self.collection, f'{folder_path_}/{filename}')
            return result_

    def to_feather(self, query=None, folder_path: str = None, filename: str = None, _id: bool = False, limit: int = -1):
        """
        pip[conda] install pyarrow
        """
        if query is None:
            query = {}
        if not isinstance(query, dict):
            raise TypeError('query must be of dict type')
        if not isinstance(limit, int):
            raise TypeError("limit must be an integer type")
        if not isinstance(_id, bool):
            raise TypeError("_id must be an boolean type")
        folder_path_ = check_folder_path(folder_path)
        if self.collection_:
            if filename is None:
                filename = f'{self.collection}_{to_str_datetime()}.feather'
            doc_objs_ = self.collection_.find(query, {"_id": 0}).limit(
                limit) if limit != -1 else self.collection_.find(query, {"_id": 0})
            doc_list_ = [schema_(doc_) for doc_ in doc_objs_]
            df_ = pa.Table.from_pylist(mapping=doc_list_)
            with open(f'{folder_path_}/{filename}', 'wb') as f:
                pa_feather_.write_feather(df_, f)

            result_ = ECHO_INFO.format(Fore.GREEN, self.collection, f'{folder_path_}/{filename}')
            return result_

    def to_parquet(self, query=None, folder_path: str = None, filename: str = None, _id: bool = False, limit: int = -1):
        if query is None:
            query = {}
        if not isinstance(query, dict):
            raise TypeError('query must be of dict type')
        if not isinstance(limit, int):
            raise TypeError("limit must be an integer type")
        if not isinstance(_id, bool):
            raise TypeError("_id must be an boolean type")
        folder_path_ = check_folder_path(folder_path)
        if self.collection_:
            if filename is None:
                filename = f'{self.collection}_{to_str_datetime()}.parquet'
            doc_objs_ = self.collection_.find(query, {"_id": 0}).limit(
                limit) if limit != -1 else self.collection_.find(query, {"_id": 0})
            doc_list_ = [schema_(doc_) for doc_ in doc_objs_]
            df_ = pa.Table.from_pylist(mapping=doc_list_)
            with open(f'{folder_path_}/{filename}', 'wb') as f:
                pa_parquet_.write_table(df_, f)
            result_ = ECHO_INFO.format(Fore.GREEN, self.collection, f'{folder_path_}/{filename}')
            return result_

    def save_csv_(self, pg: int, block_size_: int, collection_name: str, folder_path_: str, ignore_error: bool = False):
        doc_objs_ = self.collection_.find({}, {'_id': 0}, batch_size=block_size_).skip(pg * block_size_).limit(
            block_size_)
        filename = f'{collection_name}_{pg}.csv'
        # filename = f'{collection_name}_{str(uuid.uuid4())}.csv'
        doc_list_ = [schema_(doc_) for doc_ in doc_objs_]
        df_ = pa.Table.from_pylist(mapping=doc_list_, schema=None)
        with pa_csv_.CSVWriter(f'{folder_path_}/{filename}', df_.schema) as writer:
            writer.write_table(df_)

        # with codecs.open(f'{folder_path_}/{filename}', 'w', encoding=PANDAS_ENCODING) as csvfile:
        #     writer = csv.writer(csvfile)
        #     if ignore_error:
        #         ...
        #     else:
        #         writer.writerow(list(dict(doc_list_[0]).keys()))
        #         writer.writerows([list(dict(data_).values()) for data_ in doc_list_])
        # for data_ in doc_list_: writer.writerow(list(dict(data_).values()))

        return f'{Fore.GREEN} → {folder_path_}/{filename} is ok'

    def save_excel_(self, f_: xlsxwriter.Workbook, pg: int, block_size_: int, collection_name: str, folder_path_: str,
                    ignore_error: bool):
        if f_:
            work_sheet_ = f_.add_worksheet(f'Sheet{pg + 1}')
            # print(f"A{i + 1}", list(dict(doc).values()))
            doc_objs_ = self.collection_.find({}, {'_id': 0}, batch_size=block_size_).skip(pg * block_size_).limit(
                block_size_)
            header_ = list(dict(doc_objs_[0]).keys())
            work_sheet_.write_row(f"A1", header_)
            if ignore_error:
                for index_, doc_ in enumerate(doc_objs_):
                    work_sheet_.write_row(f"A{index_ + 2}",
                                          [doc_.get(_) if doc_.get(_) and isinstance(doc_.get(_), IGNORE_TYPE) else None
                                           for _ in header_])
            else:
                [work_sheet_.write_row(f"A{index_ + 2}",
                                       [doc_ if isinstance(doc_, (int, float)) or doc_ is None else str(doc_) for doc_
                                        in
                                        dict(doc).values()]) for
                 index_, doc in enumerate(doc_objs_)]
            return f'{Fore.GREEN} → {work_sheet_.name} is ok'
        else:
            filename = f'{collection_name}_{pg + 1}.xlsx'
            doc_objs_ = self.collection_.find({}, {"_id": 0}, batch_size=block_size_).skip(pg * block_size_).limit(
                block_size_)
            header_ = list(dict(doc_objs_[0]).keys())
            with xlsxwriter.Workbook(f'{folder_path_}/{filename}') as work_book_:
                work_sheet_ = work_book_.add_worksheet(f'Sheet{pg + 1}')
                work_sheet_.write_row(f"A1", header_)
                if ignore_error:
                    for index_, doc in enumerate(doc_objs_):
                        # print(list(dict(doc).values()))
                        write_list_ = [doc.get(x_) if doc.get(x_) and isinstance(doc.get(x_), IGNORE_TYPE) else None for
                                       x_ in header_]
                        work_sheet_.write_row(f"A{index_ + 2}", write_list_)
                else:
                    # work_sheet_._write_rows([list(dict(doc).values()) for i, doc in enumerate(doc_objs_)])
                    for index_, doc in enumerate(doc_objs_):
                        write_list_ = [doc_ if isinstance(doc_, (int, float)) or doc_ is None else str(doc_) for doc_ in
                                       dict(doc).values()]
                        work_sheet_.write_row(f"A{index_ + 2}", write_list_)
            return f'{Fore.GREEN} → {folder_path_}/{filename} is ok'

    def save_json_(self, pg: int, block_size_: int, collection_name: str, folder_path_: str):
        doc_objs_ = self.collection_.find({}, {'_id': 0}, batch_size=block_size_).skip(pg * block_size_).limit(
            block_size_)
        filename = f'{collection_name}_{pg}.json'
        doc_list_ = list(doc_objs_)
        data = {'RECORDS': doc_list_}
        with open(f'{folder_path_}/{filename}', 'w', encoding="utf-8") as f:
            f.write(serialize_obj(data))
        return f'{Fore.GREEN} → {folder_path_}/{filename} is ok'

    def to_many_collection_(self, func, folder_path: str, ignore_error: bool):
        concurrent_(func, self.database, self.get_collection_objs_(), folder_path, ignore_error)

    def get_collection_objs_(self):
        return [{"collection_name": collection_name, "collection_": self.db_[collection_name]} for
                collection_name in self.collection_names]


