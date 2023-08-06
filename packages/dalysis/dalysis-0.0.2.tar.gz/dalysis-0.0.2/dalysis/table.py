import csv
import prettytable
import numpy
from dalysis import typecv
import openpyxl

    
def _toFloat(variable):
    # This function is used for judge whether the varible is interger or float
    try:
        return float(variable)
    except:
        return variable

def _arrayToString(array):
    res = array.copy()
    for i in range(0,len(res)):
        res[i] = str(res[i])
    return res

class Table:
    def __init__(self,source,head=True):
        self.__source = source
        self.__table = {}
        self.__head = self.__source[0].copy()
        self.type_list = []
        if(head):
            for i in range(0,len(self.__head)):
                self.__table[ self.__head[i] ] = []
                for j in range(1,len(self.__source)):
                    self.__source[j][i] = _toFloat( self.__source[j][i] )
                    self.__table[ self.__head[i] ] .append( self.__source[j][i])
        else:
            for i in range(0,len(self.__head)):
                self.__head[i] = "col-" + str(i+1)
                self.__table[ self.__head[i] ] = []
                for j in range(1,len(self.__source)):
                    self.__source[j][i] = _toFloat( self.__source[j][i] )
                    self.__table[ self.__head[i] ] .append( self.__source[j][i])
            self.__source.insert( 0,self.__head )
        
        for i in range(0,len(self.__head)):
            self.type_list.append( type( self.__source[1][i] ) )

    def output(self):
        PrintTool = prettytable.PrettyTable()
        PrintTool.field_names = self.__head

        for i in range(1,len(self.__source)):
            PrintTool.add_row(  self.__source[i]  )
        print(PrintTool)

    def head(self,index):
        return self.__table[index]

    def row(self,index):
        return self.head( self.__head[index] )

    def col(self,index):
        return self.__source[index]

    #only add one column now
    def add_col(self,data,head=""):
        head = str(head)
        if( len(head) == 0 ):
            head = "col-" + str(len(self.__head))
        self.__head.append( head )
        # try:
        #     self.__table[head] = data
        # except:
        #     print("The dimension of new data is different to the original for row.")
        self.__table[head] = data
        for i in range( len(data) , len(self.__source) - 1 ):
            self.__table[head].append(None)
        
        self.__source[0].append( head )
        for i in range(1,len(self.__source)):
            self.__source[i].append( self.__table[head][i-1] )

        self.type_list.append(type(self.__source[1][  len(self.__head) - 1 ]))
        
    
    def add_row(self,data):
        # data_type = [ type(i) for i in data ]
        # bool_a
        # for i in range(0, len(data) ):
        #     bool_a = bool_a and ( data_type[i] == self.type_list )
        # if(bool_a):
        #     for i in range(0,len(  ))

        for i in range(0,len(data)):
            self.__table[self.__head[i]].append(data[i])
            
        self.__source.append(data)
        for i in range(len(data),len(self.__head)):
            self.__table[self.__head[i]].append(None)
            self.__source[ len(self.__source) - 1 ].append(None)


        

def open_csv(path,mod='r',head=True):
    file_source = open(path,mod)
    fcsv = csv.reader(file_source)
    listcsv = [i for i in fcsv]

    file_source.close()
    return Table(listcsv,head)


def write_csv(table_,path,mod='a',head=True):
    try:
        out = open(path,mod,newline='')
        csv_write = csv.writer(out,dialect='excel')
        for i in range(0,len( table_.row(0) )):
            csv_write.writerow( table_.col(i) )
        out.close()
        return True
    except:
        return False

def open_excel(path,sheetname="Sheet1",head=True):
    wb = openpyxl.load_workbook(path)
    ws = wb.get_sheet_by_name(sheetname)
    rows = ws.max_row
    cols = ws.max_column

    res = []
    for i in range(0,rows):
        temp = []
        for j in range(0,cols):
            temp.append( ws.cell(i+1,j+1).value )
        res.append(temp)
    
    return Table(res,head)

